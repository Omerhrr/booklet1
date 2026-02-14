"""
Report Service - Financial Reports Generation
"""

from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, case
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..models.sales import SalesInvoice
from ..models.purchase import PurchaseBill
from ..models.customer import Customer
from ..models.vendor import Vendor
from ..models.account import Account, AccountType, LedgerEntry
from .accounting_service import accounting_service


class ReportService:
    """Financial reports generation service"""
    
    # ============================================
    # AGING REPORTS
    # ============================================
    
    def get_ar_aging_report(
        self,
        db: Session,
        tenant_id: int,
        branch_id: int = None,
        as_of_date: date = None
    ) -> List[Dict]:
        """Generate Accounts Receivable aging report"""
        if not as_of_date:
            as_of_date = date.today()
        
        # Get unpaid invoices
        query = db.query(SalesInvoice).filter(
            SalesInvoice.tenant_id == tenant_id,
            SalesInvoice.status.in_(['unpaid', 'partially_paid'])
        )
        
        if branch_id:
            query = query.filter(SalesInvoice.branch_id == branch_id)
        
        invoices = query.all()
        
        # Calculate aging buckets
        aging_data = []
        
        for invoice in invoices:
            days_outstanding = (as_of_date - invoice.invoice_date).days
            balance = invoice.total_amount - invoice.paid_amount
            
            if balance <= 0:
                continue
            
            # Determine bucket
            if days_outstanding <= 30:
                bucket = 'current'
            elif days_outstanding <= 60:
                bucket = '31_60'
            elif days_outstanding <= 90:
                bucket = '61_90'
            else:
                bucket = 'over_90'
            
            # Check if customer already in aging data
            customer_entry = next(
                (a for a in aging_data if a['customer_id'] == invoice.customer_id),
                None
            )
            
            if customer_entry:
                customer_entry[bucket] += balance
                customer_entry['total'] += balance
                customer_entry['invoices'].append({
                    'invoice_number': invoice.invoice_number,
                    'date': invoice.invoice_date,
                    'due_date': invoice.due_date,
                    'amount': invoice.total_amount,
                    'paid': invoice.paid_amount,
                    'balance': balance,
                    'days_outstanding': days_outstanding
                })
            else:
                aging_data.append({
                    'customer_id': invoice.customer_id,
                    'customer_name': invoice.customer.name,
                    'customer_email': invoice.customer.email,
                    'current': balance if bucket == 'current' else 0,
                    '31_60': balance if bucket == '31_60' else 0,
                    '61_90': balance if bucket == '61_90' else 0,
                    'over_90': balance if bucket == 'over_90' else 0,
                    'total': balance,
                    'invoices': [{
                        'invoice_number': invoice.invoice_number,
                        'date': invoice.invoice_date,
                        'due_date': invoice.due_date,
                        'amount': invoice.total_amount,
                        'paid': invoice.paid_amount,
                        'balance': balance,
                        'days_outstanding': days_outstanding
                    }]
                })
        
        # Sort by total descending
        aging_data.sort(key=lambda x: x['total'], reverse=True)
        
        return aging_data
    
    def get_ap_aging_report(
        self,
        db: Session,
        tenant_id: int,
        branch_id: int = None,
        as_of_date: date = None
    ) -> List[Dict]:
        """Generate Accounts Payable aging report"""
        if not as_of_date:
            as_of_date = date.today()
        
        # Get unpaid bills
        query = db.query(PurchaseBill).filter(
            PurchaseBill.tenant_id == tenant_id,
            PurchaseBill.status.in_(['unpaid', 'partially_paid'])
        )
        
        if branch_id:
            query = query.filter(PurchaseBill.branch_id == branch_id)
        
        bills = query.all()
        
        # Calculate aging buckets
        aging_data = []
        
        for bill in bills:
            days_outstanding = (as_of_date - bill.bill_date).days
            balance = bill.total_amount - bill.paid_amount
            
            if balance <= 0:
                continue
            
            # Determine bucket
            if days_outstanding <= 30:
                bucket = 'current'
            elif days_outstanding <= 60:
                bucket = '31_60'
            elif days_outstanding <= 90:
                bucket = '61_90'
            else:
                bucket = 'over_90'
            
            # Check if vendor already in aging data
            vendor_entry = next(
                (a for a in aging_data if a['vendor_id'] == bill.vendor_id),
                None
            )
            
            if vendor_entry:
                vendor_entry[bucket] += balance
                vendor_entry['total'] += balance
                vendor_entry['bills'].append({
                    'bill_number': bill.bill_number,
                    'date': bill.bill_date,
                    'due_date': bill.due_date,
                    'amount': bill.total_amount,
                    'paid': bill.paid_amount,
                    'balance': balance,
                    'days_outstanding': days_outstanding
                })
            else:
                aging_data.append({
                    'vendor_id': bill.vendor_id,
                    'vendor_name': bill.vendor.name,
                    'vendor_email': bill.vendor.email,
                    'current': balance if bucket == 'current' else 0,
                    '31_60': balance if bucket == '31_60' else 0,
                    '61_90': balance if bucket == '61_90' else 0,
                    'over_90': balance if bucket == 'over_90' else 0,
                    'total': balance,
                    'bills': [{
                        'bill_number': bill.bill_number,
                        'date': bill.bill_date,
                        'due_date': bill.due_date,
                        'amount': bill.total_amount,
                        'paid': bill.paid_amount,
                        'balance': balance,
                        'days_outstanding': days_outstanding
                    }]
                })
        
        # Sort by total descending
        aging_data.sort(key=lambda x: x['total'], reverse=True)
        
        return aging_data
    
    # ============================================
    # SALES REPORTS
    # ============================================
    
    def get_sales_report(
        self,
        db: Session,
        tenant_id: int,
        branch_id: int = None,
        start_date: date = None,
        end_date: date = None
    ) -> Dict:
        """Generate sales report"""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        query = db.query(SalesInvoice).filter(
            SalesInvoice.tenant_id == tenant_id,
            SalesInvoice.invoice_date >= start_date,
            SalesInvoice.invoice_date <= end_date
        )
        
        if branch_id:
            query = query.filter(SalesInvoice.branch_id == branch_id)
        
        invoices = query.order_by(SalesInvoice.invoice_date).all()
        
        # Calculate totals
        total_subtotal = sum(i.subtotal for i in invoices)
        total_vat = sum(i.vat_amount for i in invoices)
        total_amount = sum(i.total_amount for i in invoices)
        total_paid = sum(i.paid_amount for i in invoices)
        total_outstanding = total_amount - total_paid
        
        # Group by customer
        by_customer = {}
        for invoice in invoices:
            customer_name = invoice.customer.name
            if customer_name not in by_customer:
                by_customer[customer_name] = {
                    'customer_name': customer_name,
                    'invoice_count': 0,
                    'total_amount': 0
                }
            by_customer[customer_name]['invoice_count'] += 1
            by_customer[customer_name]['total_amount'] += invoice.total_amount
        
        # Group by month
        by_month = {}
        for invoice in invoices:
            month_key = invoice.invoice_date.strftime('%Y-%m')
            if month_key not in by_month:
                by_month[month_key] = {
                    'month': month_key,
                    'invoice_count': 0,
                    'total_amount': 0
                }
            by_month[month_key]['invoice_count'] += 1
            by_month[month_key]['total_amount'] += invoice.total_amount
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'total_invoices': len(invoices),
            'total_subtotal': total_subtotal,
            'total_vat': total_vat,
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_outstanding': total_outstanding,
            'by_customer': list(by_customer.values()),
            'by_month': list(by_month.values()),
            'invoices': invoices
        }
    
    # ============================================
    # PURCHASE REPORTS
    # ============================================
    
    def get_purchase_report(
        self,
        db: Session,
        tenant_id: int,
        branch_id: int = None,
        start_date: date = None,
        end_date: date = None
    ) -> Dict:
        """Generate purchase report"""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        query = db.query(PurchaseBill).filter(
            PurchaseBill.tenant_id == tenant_id,
            PurchaseBill.bill_date >= start_date,
            PurchaseBill.bill_date <= end_date
        )
        
        if branch_id:
            query = query.filter(PurchaseBill.branch_id == branch_id)
        
        bills = query.order_by(PurchaseBill.bill_date).all()
        
        # Calculate totals
        total_subtotal = sum(b.subtotal for b in bills)
        total_vat = sum(b.vat_amount for b in bills)
        total_amount = sum(b.total_amount for b in bills)
        total_paid = sum(b.paid_amount for b in bills)
        total_outstanding = total_amount - total_paid
        
        # Group by vendor
        by_vendor = {}
        for bill in bills:
            vendor_name = bill.vendor.name
            if vendor_name not in by_vendor:
                by_vendor[vendor_name] = {
                    'vendor_name': vendor_name,
                    'bill_count': 0,
                    'total_amount': 0
                }
            by_vendor[vendor_name]['bill_count'] += 1
            by_vendor[vendor_name]['total_amount'] += bill.total_amount
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'total_bills': len(bills),
            'total_subtotal': total_subtotal,
            'total_vat': total_vat,
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_outstanding': total_outstanding,
            'by_vendor': list(by_vendor.values()),
            'bills': bills
        }
    
    # ============================================
    # DASHBOARD DATA
    # ============================================
    
    def get_dashboard_kpis(
        self,
        db: Session,
        tenant_id: int,
        branch_id: int = None,
        start_date: date = None,
        end_date: date = None
    ) -> Dict:
        """Get dashboard KPIs"""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = date(end_date.year, end_date.month, 1)  # Start of month
        
        # Sales
        sales_query = db.query(func.sum(SalesInvoice.total_amount)).filter(
            SalesInvoice.tenant_id == tenant_id,
            SalesInvoice.invoice_date >= start_date,
            SalesInvoice.invoice_date <= end_date
        )
        if branch_id:
            sales_query = sales_query.filter(SalesInvoice.branch_id == branch_id)
        total_sales = float(sales_query.scalar() or 0)
        
        # Purchases
        purchase_query = db.query(func.sum(PurchaseBill.total_amount)).filter(
            PurchaseBill.tenant_id == tenant_id,
            PurchaseBill.bill_date >= start_date,
            PurchaseBill.bill_date <= end_date
        )
        if branch_id:
            purchase_query = purchase_query.filter(PurchaseBill.branch_id == branch_id)
        total_purchases = float(purchase_query.scalar() or 0)
        
        # Receivables (AR)
        ar_query = db.query(
            func.sum(SalesInvoice.total_amount - SalesInvoice.paid_amount)
        ).filter(
            SalesInvoice.tenant_id == tenant_id,
            SalesInvoice.status.in_(['unpaid', 'partially_paid'])
        )
        if branch_id:
            ar_query = ar_query.filter(SalesInvoice.branch_id == branch_id)
        total_receivables = float(ar_query.scalar() or 0)
        
        # Payables (AP)
        ap_query = db.query(
            func.sum(PurchaseBill.total_amount - PurchaseBill.paid_amount)
        ).filter(
            PurchaseBill.tenant_id == tenant_id,
            PurchaseBill.status.in_(['unpaid', 'partially_paid'])
        )
        if branch_id:
            ap_query = ap_query.filter(PurchaseBill.branch_id == branch_id)
        total_payables = float(ap_query.scalar() or 0)
        
        # Get P&L
        pnl = accounting_service.get_profit_and_loss(
            db, tenant_id, branch_id, start_date, end_date
        )
        
        # Get balance sheet
        bs = accounting_service.get_balance_sheet(
            db, tenant_id, branch_id, end_date
        )
        
        # Cash/Bank balance
        bank_balance = 0
        cash_balance = 0
        for asset in bs['assets']['current']:
            if 'Bank' in asset['name']:
                bank_balance += asset['amount']
            elif 'Cash' in asset['name']:
                cash_balance += asset['amount']
        
        return {
            'period': {
                'start': start_date,
                'end': end_date
            },
            'sales': total_sales,
            'purchases': total_purchases,
            'receivables': total_receivables,
            'payables': total_payables,
            'gross_profit': pnl['gross_profit'],
            'net_profit': pnl['net_profit'],
            'bank_balance': bank_balance,
            'cash_balance': cash_balance,
            'total_assets': bs['total_assets']
        }
    
    # ============================================
    # EXPORT TO EXCEL
    # ============================================
    
    def export_to_excel(self, data: List[Dict], title: str, headers: List[str]) -> io.BytesIO:
        """Export data to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:' + get_column_letter(len(headers)) + '1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                # Map header to data key
                key = header.lower().replace(' ', '_').replace('(₦)', '').strip()
                value = row_data.get(key, '')
                
                # Format currency
                if isinstance(value, float) and 'amount' in key or 'total' in key or 'balance' in key:
                    value = f"₦{value:,.2f}"
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


# Singleton instance
report_service = ReportService()
